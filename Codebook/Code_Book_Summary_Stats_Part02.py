import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

# Function to return the new static path. Insert pathname
def get_path():
    return r'pathname'

# Function to replace -9 and -99 with NaN for summary stats, but keep the overall count intact
def replace_missing_values(df, variables, missing_values=[-9, -99]):
    for var in variables:
        df[var].replace(missing_values, np.nan, inplace=True)

# Summarize the data, preserving the count and calculating statistics only for valid values
def summarize_variable(df, var, is_categorical=False):
    if pd.api.types.is_numeric_dtype(df[var]):
        # Step 1: Preserve the count (including -9 and -99 as valid values)
        total_count = df[var].notna().sum()

        # Step 2: Count how many -9 and -99 are present for missing
        missing_count = (df[var] == -9).sum() + (df[var] == -99).sum()

        # Step 3: Handle categorical variables separately by removing value '2'
        if is_categorical:
            valid_values = df[var].replace([-9, -99, 2], np.nan)
        else:
            valid_values = df[var].replace([-9, -99], np.nan)

        # Step 4: Calculate summary statistics
        summary_stats = {
            'Count': total_count,  # Preserved count with -9 and -99
            'Min': valid_values.min(),
            'Mean': valid_values.mean(),
            'Max': valid_values.max(),
            'StdDev': valid_values.std(),
            'Missing': missing_count  # Count of -9 and -99 as missing
        }

        # Create DataFrame to align columns
        summary_df = pd.DataFrame([summary_stats], index=[var])
        summary_df = summary_df[['Count', 'Min', 'Mean', 'Max', 'StdDev', 'Missing']]
        return summary_df
    else:
        print(f"Skipping variable {var} because it is not numeric.")
        return None

# Function to generate frequency table with bold labels and proportions, with label before the frequency and proportion
def tabulate_variable(df, var):
    frequency_table = df[var].value_counts(normalize=False)
    proportion_table = df[var].value_counts(normalize=True).round(2)

    # Create DataFrame for tabulation
    tabulation = pd.DataFrame({
        'Label': frequency_table.index.astype(str) + ":",
        'Frequency': frequency_table.values,
        'Proportion': proportion_table.values
    })
    
    # Add the variable name at the top of the table for clarity
    tabulation_with_varname = pd.DataFrame({f'Tabulation of {var}': [f'Tabulation of {var}']})
    
    # Concatenate the label row and the tabulation data
    tabulation_with_varname = pd.concat([tabulation_with_varname, tabulation], ignore_index=False)
    
    return tabulation_with_varname

# Load data function
def load_data(path, filename):
    return pd.read_stata(f"{path}\\{filename}.dta")

# Function to export data to Excel
def export_to_excel(summary_data, tabulation_data, output_file):
    # Create a workbook and add a worksheet
    wb = Workbook()
    
    # Write summary stats to a sheet
    ws_summary = wb.active
    ws_summary.title = "Summary Statistics"
    
    for df in summary_data:
        if df is not None:
            for r in dataframe_to_rows(df, index=True, header=True):
                ws_summary.append(r)
    
    # Write tabulations to a separate sheet
    ws_tab = wb.create_sheet(title="Tabulations")
    bold_font = Font(bold=True)
    
    for df in tabulation_data:
        for r in dataframe_to_rows(df, index=False, header=True):
            ws_tab.append(r)
    
        # Apply bold font to the "Label" column (index 1)
        for row in ws_tab.iter_rows(min_row=2, max_row=ws_tab.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.font = bold_font
    
    # Save workbook to file
    wb.save(output_file)

# Main function to process specified variables for summarization and tabulation
def main():
    # Set path
    path = get_path()

    # Load specific data files based on your request
    income_data = load_data(path, 'Complete_Baseline_Income')
    living_data = load_data(path, 'Complete_Baseline_Standard_Of_Living')
    beliefs_data = load_data(path, 'Complete_Baseline_Beliefs')
    donations_data = load_data(path, 'Complete_Baseline_Public_Goods_Game')
    enum_data = load_data(path, 'Complete_Baseline_Enumerator_Observations')
    community_data = load_data(path, 'Complete_Baseline_Community')

    # Containers for summary stats and tabulations
    summary_data = []
    tabulation_data = []

    # Replace missing values for income variables (-9, -99)
    income_vars = [
        'agri_income_01', 'agri_income_03', 'agri_income_04', 'agri_income_05', 'agri_income_06',
        'agri_income_07_1', 'agri_income_07_2', 'agri_income_07_3', 'agri_income_07_4', 'agri_income_07_5', 'agri_income_07_6', 'agri_income_07_o',
        'agri_income_08_1', 'agri_income_08_2', 'agri_income_08_3', 'agri_income_08_4', 'agri_income_08_5', 'agri_income_08_6', 'agri_income_08_o',
        'agri_income_10_1', 'agri_income_10_2', 'agri_income_10_3', 'agri_income_10_4', 'agri_income_10_5', 'agri_income_10_6', 'agri_income_10_o',
        'animals_sales_1', 'animals_sales_2', 'animals_sales_3', 'animals_sales_4', 'animals_sales_5', 'animals_sales_6', 'animals_sales_7', 'animals_sales_8', 'animals_sales_9', 'animals_sales_o', 'animals_sales_t',
        'species_1', 'species_2', 'species_3', 'species_4', 'species_5', 'species_6', 'species_7', 'species_8', 'species_9', 'species_autre',
        'agri_income_11_1', 'agri_income_11_2', 'agri_income_11_3', 'agri_income_11_4', 'agri_income_11_5', 'agri_income_11_o',
        'agri_income_12_1', 'agri_income_12_2', 'agri_income_12_3', 'agri_income_12_4', 'agri_income_12_5', 'agri_income_12_o',
        'agri_income_14_1', 'agri_income_14_2', 'agri_income_14_3', 'agri_income_14_4', 'agri_income_14_5', 'agri_income_14_o',
        'agri_income_15', 'agri_income_16', 'agri_income_17', 'agri_income_18', 'agri_income_19',
        'agri_income_20_1', 'agri_income_20_2', 'agri_income_20_3', 'agri_income_20_4', 'agri_income_20_5', 'agri_income_20_6', 'agri_income_20_7', 'agri_income_20_8', 'agri_income_20_9', 'agri_income_20_t', 'agri_income_20_o',
        'agri_income_22_1', 'agri_income_22_2', 'agri_income_22_3', 'agri_income_22_4', 'agri_income_22_o',
        'agri_income_23_1', 'agri_income_23_2', 'agri_income_23_3', 'agri_income_23_4', 'agri_income_23_o',
        'agri_income_24_1', 'agri_income_24_2', 'agri_income_24_3', 'agri_income_24_4',
        'agri_income_25', 'agri_income_26', 'agri_income_27', 'agri_income_28', 'agri_income_29', 'agri_income_30',
        'agri_income_31_1', 'agri_income_31_2', 'agri_income_31_3', 'agri_income_31_4', 'agri_income_31_5', 'agri_income_31_6', 'agri_income_31_o',
        'agri_income_32', 'agri_income_33', 'agri_income_34', 'agri_income_35',
        'agri_income_36_1', 'agri_income_36_2', 'agri_income_36_3', 'agri_income_36_4',
        'agri_income_38_1', 'agri_income_38_2', 'agri_income_38_3', 'agri_income_38_4',
        'agri_income_39_1', 'agri_income_39_2', 'agri_income_39_3', 'agri_income_39_4',
        'agri_income_40', 'agri_income_41_1', 'agri_income_41_2',
        'agri_income_42_1', 'agri_income_42_2',
        'agri_income_43_1', 'agri_income_43_2'
    ]
    
    # Categorical variables that require dropping 2's
    categorical_vars = [
        'agri_income_01', 'animals_sales_o', 'species_autre', 'agri_income_15', 'agri_income_17',
        'agri_income_20_t', 'agri_income_25', 'agri_income_27', 'agri_income_30',
        'agri_income_32', 'agri_income_34', 'agri_income_40'
    ]
    
    # Process each variable
    for var in income_vars:
        if var in categorical_vars:
            summary = summarize_variable(income_data, var, is_categorical=True)
        else:
            summary = summarize_variable(income_data, var)
        
        if summary is not None:
            summary_data.append(summary)

    # Summarize living_02 and tabulate living_01 to living_06
    summary_data.append(summarize_variable(living_data, 'living_02'))
    living_vars = ['living_01', 'living_03', 'living_04', 'living_05', 'living_06']
    for var in living_vars:
        tabulation_data.append(tabulate_variable(living_data, var))

    # Tabulate beliefs_01 to beliefs_09
    for i in range(1, 10):
        tabulation_data.append(tabulate_variable(beliefs_data, f'beliefs_0{i}'))

    # Tabulate additional requested variables
    tabulation_data.append(tabulate_variable(income_data, 'agri_income_01'))
    for i in range(1, 7):
        tabulation_data.append(tabulate_variable(income_data, f'agri_income_09_{i}'))
    tabulation_data.append(tabulate_variable(income_data, 'agri_income_18'))
    tabulation_data.append(tabulate_variable(income_data, 'agri_income_35'))

    # Summarize donations data (explicit format)
    for var in donations_data.columns:
        summary = summarize_variable(donations_data, var)
        if summary is not None:
            summary_data.append(summary)

    # Summarize enum_01, enum_02 and tabulate enum_03 to enum_08
    summary_data.append(summarize_variable(enum_data, 'enum_01'))
    summary_data.append(summarize_variable(enum_data, 'enum_02'))
    enum_vars = ['enum_03', 'enum_04', 'enum_05', 'enum_06', 'enum_08']
    for var in enum_vars:
        tabulation_data.append(tabulate_variable(enum_data, var))

  # Summarize and tabulate community data

    # Community summary variables (replacing negatives with NaN but preserving count)
    community_summary_vars = [
        'q_16', 'q_17', 'q_18', 'q_19', 'q_20', 'q_21', 'q_22', 'q_23', 'q_24', 'q_25',
        'q_26', 'q_27', 'q_28', 'q_28a', 'q_29', 'q_29a', 'q_30', 'q_30a', 'q_31', 'q_31a',
        'q_32', 'q_32a', 'q_33', 'q_33a', 'q_34', 'q_35_check', 'q_52', 'q_53', 'q_54', 'q_55',
        'q_57', 'q_58', 'q60', 'q61', 'q63_1', 'q63_2', 'q63_3', 'q63_4', 'q63_5', 'q63_6', 'q63_7', 'q63_8', 'q63_9', 'q63_10', 'q64', 'q65', 'q66'
    ]

    # Categorical variables in community data that require dropping 2's
    community_categorical_vars = [
        'q_16', 'q_17', 'q_18', 'q_19', 'q_20', 'q_21', 'q_22', 'q_23', 'q_24', 'q_25', 
        'q_26', 'q_27', 'q_28', 'q_29', 'q_30', 'q_31', 'q_32', 'q_33', 'q_34', 'q_35_check',
        'q_36', 'q_37', 'q_38', 'q_39', 'q_41'
    ]

    # Process each variable for community data
    for var in community_summary_vars:
        if var in community_categorical_vars:
            summary = summarize_variable(community_data, var, is_categorical=True)
        else:
            summary = summarize_variable(community_data, var)
        
        if summary is not None:
            summary_data.append(summary)

    # Tabulate q62
    tabulation_data.append(tabulate_variable(community_data, 'q62'))

    # Export to Excel insert pathname
    output_file = r'pathname\DISES_Codebook_Summary_Statistics_and_Tabulations.xlsx'
    export_to_excel(summary_data, tabulation_data, output_file)

    print(f"Summary statistics and tabulations exported to {output_file}")

# Call main function
if __name__ == "__main__":
    main()
